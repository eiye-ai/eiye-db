"""Document extractors shared by the filesystem and S3 connectors.

Both connectors expose the same four shapes — CSV as rows, XLSX as rows, PDF
and text as one `content` row — and differ only in where the bytes come from.
So the parsing lives here and each connector supplies an open stream: a file on
disk, or an object body already read into memory.

Nothing here touches a path or a bucket. Callers wrap the errors with whatever
names the thing that failed, because "cannot read receipt.pdf" is useful and
"cannot read <_io.BytesIO object>" is not.
"""

import csv
from typing import IO, Any

from openpyxl import load_workbook
from pypdf import PdfReader

from eiye_db.connectors.base import ConnectorError

TEXT_SUFFIXES = {".txt", ".md", ".log", ".json"}
SAMPLE_ROWS = 20
MAX_TEXT_CHARS = 100_000
MAX_XLSX_COLS = 256
MAX_CELL_CHARS = 10_000


def kind_for(name: str) -> str | None:
    """Name the extractor for a file name or object key; None if there is none.

    Suffix-driven, like the filesystem connector has always been: the bytes are
    never sniffed, so a `.csv` holding JSON is a parse error rather than a
    silent reinterpretation.
    """
    lowered = name.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx"):
        return "xlsx"
    if lowered.endswith(".pdf"):
        return "pdf"
    if any(lowered.endswith(s) for s in TEXT_SUFFIXES):
        return "text"
    return None


def infer_type(values: list[str]) -> str:
    non_empty = [v for v in values if v != ""]
    if not non_empty:
        return "string"
    try:
        for v in non_empty:
            int(v)
        return "integer"
    except ValueError:
        pass
    try:
        for v in non_empty:
            float(v)
        return "number"
    except ValueError:
        return "string"


def csv_fields(stream: IO[str]) -> list[dict[str, Any]]:
    """Header names plus a type inferred from the first SAMPLE_ROWS rows.

    `stream` must have been opened with `newline=""`, per the csv module.
    """
    reader = csv.reader(stream)
    header = next(reader, None)
    if not header:
        return []
    samples: list[list[str]] = [[] for _ in header]
    for i, row in enumerate(reader):
        if i >= SAMPLE_ROWS:
            break
        for col, value in enumerate(row[: len(header)]):
            samples[col].append(value)
    return [{"name": h, "type": infer_type(samples[i])} for i, h in enumerate(header)]


def csv_rows(stream: IO[str], limit: int) -> list[dict[str, Any]]:
    # restkey keeps overflow cells from ragged rows under a string key instead
    # of None (which breaks JSON serialization).
    reader = csv.DictReader(stream, restkey="_extra")
    return [row for _, row in zip(range(limit), reader)]


def xlsx_fields(source: str | IO[bytes]) -> list[dict[str, Any]]:
    """Best-effort field inference from the active sheet's header + sample rows.

    Returns [] rather than raising if the workbook cannot be read: one corrupt
    spreadsheet must not fail discovery for the whole datasource.
    """
    try:
        wb = load_workbook(filename=source, read_only=True, data_only=True)
    except Exception:
        return []
    try:
        ws = wb.active
        if ws is None:
            return []
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        cols = _header_columns(header)
        samples: list[list[str]] = [[] for _ in cols]
        for i, row in enumerate(rows):
            if i >= SAMPLE_ROWS:
                break
            for c, value in enumerate(row[: len(cols)]):
                samples[c].append(_cell(value))
        return [{"name": cols[i], "type": infer_type(samples[i])} for i in range(len(cols))]
    except Exception:
        return []
    finally:
        wb.close()


def xlsx_rows(source: str | IO[bytes], limit: int, label: str) -> list[dict[str, Any]]:
    """Rows from the active sheet. Multi-sheet workbooks expose only that one."""
    try:
        wb = load_workbook(filename=source, read_only=True, data_only=True)
    except Exception as e:
        raise ConnectorError(f"cannot read xlsx {label}: {e}") from e
    try:
        ws = wb.active
        if ws is None:
            return []
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        cols = _header_columns(header)
        out: list[dict[str, Any]] = []
        for row in rows:
            if len(out) >= limit:
                break
            out.append({cols[i]: _cell(v) for i, v in enumerate(row[: len(cols)])})
        return out
    except ConnectorError:
        raise
    except Exception as e:
        raise ConnectorError(f"cannot read xlsx {label}: {e}") from e
    finally:
        wb.close()


def _header_columns(header) -> list[str]:
    return [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)][:MAX_XLSX_COLS]


def _cell(value) -> str:
    return "" if value is None else str(value)[:MAX_CELL_CHARS]


def pdf_rows(source: str | IO[bytes], label: str) -> list[dict[str, Any]]:
    try:
        reader = PdfReader(source)
    except Exception as e:
        raise ConnectorError(f"cannot read PDF {label}: {e}") from e
    if reader.is_encrypted:
        raise ConnectorError(f"encrypted PDF not supported: {label}")
    parts: list[str] = []
    total = 0
    try:
        for page in reader.pages:
            text = page.extract_text() or ""
            parts.append(text)
            total += len(text)
            if total >= MAX_TEXT_CHARS:  # stop before pulling a huge doc fully into memory
                break
    except Exception as e:
        raise ConnectorError(f"cannot extract text from PDF {label}: {e}") from e
    return [{"content": "\n".join(parts)[:MAX_TEXT_CHARS]}]


def text_rows(text: str) -> list[dict[str, Any]]:
    return [{"content": text[:MAX_TEXT_CHARS]}]
